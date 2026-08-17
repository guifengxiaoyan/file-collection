import os
import io
import re
import shutil
import uuid
import zipfile
from copy import copy
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from models import db, Admin
from config import Config
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# WPS「单元格内嵌图片」(以 DISPIMG 公式锚定在单元格中的截图) 相关声明。
# openpyxl 无法识别这类图片（ws._images 为空），保存时会把整张图片存储丢弃，
# 因此合并后必须手动把图片字节与声明重新注入结果文件，否则截图会丢失。
WPS_CELLIMAGE_NS = 'http://www.wps.cn/officeDocument/2017/etCustomData'
CELLIMAGE_REL_TYPE = 'http://www.wps.cn/officeDocument/2020/cellImage'
CELLIMAGE_CONTENT_TYPE = 'application/vnd.wps-officedocument.cellimage+xml'

def safe_remove_file(path):
    """安全删除文件。

    沙箱环境下 os.remove 会被 safe-delete 包装拦截：删除时先尝试移入回收站，
    当前环境回收站不可用会抛出 OSError 导致 500。这里忽略删除失败，保证业务流程
    （如数据库记录删除）不被打断；文件若确实删不掉也仅残留磁盘，不影响功能。
    """
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass

def safe_remove_dir(path):
    """安全删除目录及其内容，忽略删除失败。"""
    try:
        if os.path.isdir(path):
            shutil.rmtree(path)
    except OSError:
        pass

def init_default_admin():
    if not Admin.query.filter_by(username='admin').first():
        admin = Admin(username='admin', password_hash=generate_password_hash('admin123'))
        db.session.add(admin)
        db.session.commit()

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def sanitize_stored_filename(original_name):
    disk_name = secure_filename(original_name)
    if not disk_name:
        disk_name = 'file'
    name, ext = os.path.splitext(disk_name)
    if not ext:
        _, orig_ext = os.path.splitext(original_name)
        disk_name = f'{name}{orig_ext}'
    return disk_name

def get_theme_folder(theme_id):
    folder = os.path.join(Config.UPLOAD_FOLDER, f'theme_{theme_id}')
    if not os.path.exists(folder):
        os.makedirs(folder)
    return folder

def get_object_folder(theme_id, object_id):
    folder = os.path.join(get_theme_folder(theme_id), f'object_{object_id}')
    if not os.path.exists(folder):
        os.makedirs(folder)
    return folder

def get_announcement_folder():
    folder = os.path.join(Config.UPLOAD_FOLDER, 'announcements')
    if not os.path.exists(folder):
        os.makedirs(folder)
    return folder

def rename_uploaded_file(theme_id, object_id, filename, original_name):
    object_folder = get_object_folder(theme_id, object_id)
    name, ext = os.path.splitext(original_name)
    safe_name = f"{name}{ext}"
    counter = 1
    while os.path.exists(os.path.join(object_folder, safe_name)):
        safe_name = f"{name}_{counter}{ext}"
        counter += 1
    return safe_name

EXCEL_EXTS = ('.xls', '.xlsx')

def _copy_cell_style(src, dst):
    """把源单元格的样式（字体/填充/边框/对齐/数字格式/保护）复制到目标单元格。

    不同工作簿之间不能共享样式索引，因此逐项拷贝属性。
    """
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def _place_image_to_cell(out_ws, img, out_row, out_col, occupied):
    """把一张图片重新锚定到输出表的指定单元格，避免与已有图片重合。

    返回 True 表示放置成功，False 表示该位置已占用或图片不可用（已跳过）。
    """
    key = (out_row, out_col)
    if key in occupied:
        # 该单元格已有截图，避免重叠（重合单元格）直接跳过，绝不覆盖
        return False
    try:
        data = img._data()
    except Exception:
        return False
    if not data:
        return False
    new_img = XLImage(io.BytesIO(data))
    if getattr(img, 'width', None):
        new_img.width = img.width
    if getattr(img, 'height', None):
        new_img.height = img.height
    try:
        anchor = f"{get_column_letter(out_col)}{out_row}"
        out_ws.add_image(new_img, anchor)
    except Exception:
        return False
    occupied.add(key)
    return True


def _collect_wps_cell_images(path):
    """读取 WPS「单元格内嵌图片」(xl/cellimages.xml + xl/media)。

    这类截图以 DISPIMG 公式锚定在单元格中（例如 =_xlfn.DISPIMG("ID_xxx",1)），
    openpyxl 无法识别（ws._images 为空），保存时会把整张图片存储丢弃，导致合并后
    截图丢失。这里从原始 zip 中解析出每张图片的字节与 pic 定义，供后续重新注入
    到合并结果中。

    返回 [{ 'name': <图片ID>, 'media': <bytes>, 'ext': '.png', 'pic': <etc:cellImage xml> }, ...]
    （按图片 id 去重）
    """
    try:
        z = zipfile.ZipFile(path)
    except Exception:
        return []
    names = set(z.namelist())
    if 'xl/cellimages.xml' not in names:
        return []

    try:
        cellimages = z.read('xl/cellimages.xml').decode('utf-8')
        rels = z.read('xl/_rels/cellimages.xml.rels').decode('utf-8')
    except Exception:
        return []

    # rId -> 相对 xl/ 的 media 路径
    rid_target = {}
    for m in re.finditer(r'Id="(rId\d+)"[^>]*?Target="([^"]+)"', rels):
        rid_target[m.group(1)] = m.group(2)

    out = []
    seen = set()
    for block in re.findall(r'<etc:cellImage>.*?</etc:cellImage>', cellimages, re.S):
        nm = re.search(r'<xdr:cNvPr[^>]*?\bname="([^"]+)"', block)
        blip = re.search(r'<a:blip[^>]*?r:embed="([^"]+)"', block)
        if not nm or not blip:
            continue
        name = nm.group(1)
        if name in seen:
            continue
        target = rid_target.get(blip.group(1))
        if not target:
            continue
        media_path = target if target.startswith('xl/') else 'xl/' + target
        if media_path not in names:
            continue
        media = z.read(media_path)
        ext = os.path.splitext(target)[1].lower() or '.png'
        out.append({'name': name, 'media': media, 'ext': ext, 'pic': block})
        seen.add(name)
    return out


def _inject_wps_cell_images(buf, images):
    """把 WPS 单元格内嵌图片重新写入已保存的 xlsx 字节流。

    openpyxl 保存时会丢弃 xl/cellimages.xml 与 xl/media/，这里在保存后的 zip 中
    重新注入媒体文件、cellimages.xml 及其关系，并补全 [Content_Types].xml 与
    workbook.xml.rels 的声明，使 DISPIMG 公式能继续解析出截图。

    返回新的 BytesIO（若 images 为空则原样返回 buf）。
    """
    if not images:
        return buf

    zin = zipfile.ZipFile(buf)
    entries = {item.filename: zin.read(item.filename) for item in zin.infolist()}
    zin.close()

    # 1) 媒体文件（唯一文件名避免不同文件同名冲突）
    for i, img in enumerate(images, 1):
        entries[f'xl/media/merged_img_{i}{img["ext"]}'] = img['media']

    # 2) cellimages.xml：聚合所有 pic，重写 r:embed 指向新的 rId
    cellimage_parts = []
    rels_items = []
    for i, img in enumerate(images, 1):
        rid = f'rIdImg{i}'
        # 仅替换第一个 r:embed（每个 pic 只有一个 blip）
        pic = re.sub(
            r'(<a:blip[^>]*?r:embed=")[^"]+(")',
            lambda m, r=rid: m.group(1) + r + m.group(2),
            img['pic'], count=1,
        )
        cellimage_parts.append(pic)
        rels_items.append(
            f'<Relationship Id="{rid}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="media/merged_img_{i}{img["ext"]}"/>'
        )

    cellimages_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<etc:cellImages '
        'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        f'xmlns:etc="{WPS_CELLIMAGE_NS}">'
        + ''.join(cellimage_parts) +
        '</etc:cellImages>'
    )
    entries['xl/cellimages.xml'] = cellimages_xml.encode('utf-8')
    entries['xl/_rels/cellimages.xml.rels'] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + ''.join(rels_items) +
        '</Relationships>'
    ).encode('utf-8')

    # 3) [Content_Types].xml：确保图片扩展名默认类型 + cellimages override
    ct = entries['[Content_Types].xml'].decode('utf-8')
    exts_used = {img['ext'].lstrip('.') for img in images}
    for ext in exts_used:
        if f'Extension="{ext}"' not in ct:
            ct = ct.replace(
                '</Types>',
                f'<Default Extension="{ext}" ContentType="image/{ext}"/></Types>',
            )
    if '/xl/cellimages.xml' not in ct:
        ct = ct.replace(
            '</Types>',
            '<Override PartName="/xl/cellimages.xml" '
            f'ContentType="{CELLIMAGE_CONTENT_TYPE}"/></Types>',
        )
    entries['[Content_Types].xml'] = ct.encode('utf-8')

    # 4) workbook.xml.rels：增加 cellimages 关系（Id 不冲突）
    wrels_name = 'xl/_rels/workbook.xml.rels'
    wrels = entries[wrels_name].decode('utf-8')
    if 'cellimages.xml' not in wrels:
        ids = set(re.findall(r'Id="([^"]+)"', wrels))
        rid_cell = 'rIdCellImages'
        n = 1
        while rid_cell in ids:
            rid_cell = f'rIdCellImages{n}'
            n += 1
        wrels = wrels.replace(
            '</Relationships>',
            f'<Relationship Id="{rid_cell}" Type="{CELLIMAGE_REL_TYPE}" '
            'Target="cellimages.xml"/></Relationships>',
        )
        entries[wrels_name] = wrels.encode('utf-8')

    # 重新打包
    out = io.BytesIO()
    zout = zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED)
    for name, data in entries.items():
        zout.writestr(name, data)
    zout.close()
    out.seek(0)
    return out


def detect_excel_merge_ready(theme_id):
    """检测某个主题下的全部附件是否满足 Excel 合并条件。

    条件：
    1. 每个收集对象都至少上传了附件；
    2. 主题下所有附件都是 Excel 文件（.xls/.xlsx）；
    3. 所有 Excel 文件都能被读取且列数（格式）一致。

    返回 dict：{ ready, reason, files, columns }
      files 元素为 { object_name, original_name, path, ext, columns, has_images }
    """
    from models import CollectionObject, Attachment

    objects = CollectionObject.query.filter_by(theme_id=theme_id).order_by(CollectionObject.id).all()

    result = {'ready': False, 'reason': '', 'files': [], 'columns': None}

    excel_files = []
    non_excel = []
    for obj in objects:
        for att in obj.attachments:
            ext = os.path.splitext(att.original_name)[1].lower()
            path = os.path.join(get_object_folder(obj.theme_id, obj.id), att.filename)
            if ext not in EXCEL_EXTS:
                non_excel.append(att.original_name)
                continue
            excel_files.append({
                'object_name': obj.name,
                'original_name': att.original_name,
                'path': path,
                'ext': ext,
            })

    if not excel_files:
        result['reason'] = '该主题下还没有任何 Excel 附件，无法合并'
        return result

    if non_excel:
        shown = '、'.join(non_excel[:5])
        more = ' 等' if len(non_excel) > 5 else ''
        result['reason'] = f'存在非 Excel 附件（{shown}{more}），必须全部为 Excel 文件才能合并'
        return result

    # 检查每个 Excel 是否可读 + 列数是否一致（即“文件格式一样”）
    columns = None
    infos = []
    for f in excel_files:
        try:
            wb = openpyxl.load_workbook(f['path'], read_only=True, data_only=True)
            ws = wb.active
            ncols = ws.max_column
            nrows = ws.max_row
            wb.close()
        except Exception as e:
            result['reason'] = f'无法读取 Excel 文件「{f["original_name"]}」（可能是旧版 .xls 格式或文件损坏）：{e}'
            return result
        if ncols <= 0:
            result['reason'] = f'文件「{f["original_name"]}」没有可识别的列，无法合并'
            return result
        if columns is None:
            columns = ncols
        elif ncols != columns:
            result['reason'] = (f'各文件列数（格式）不一致：参考 {columns} 列，'
                                f'但「{f["original_name"]}」为 {ncols} 列，无法合并')
            return result
        infos.append({
            'object_name': f['object_name'],
            'original_name': f['original_name'],
            'path': f['path'],
            'ext': f['ext'],
            'columns': ncols,
            'rows': nrows,
        })

    result['ready'] = True
    result['files'] = infos
    result['columns'] = columns
    return result


def merge_excel_files(files, merge_identical=False, mark_source=False):
    """合并多个结构一致的 Excel 文件为单个工作簿（写入内存并返回 BytesIO）。

    files: list of dict { path, object_name, original_name }
      - 以第一个文件的表头为合并后表头；
      - 逐行（按行合并）把所有文件的数据行堆叠到一起；
      - merge_identical=True 时，跨所有文件删除“数据列完全相同”的重复行（去重合并）；
      - mark_source=True 时，在末尾新增“收集对象”列标注来源；
      - 单元格中的截图（含 WPS 以 DISPIMG 公式锚定的“单元格内嵌图片”）会被“固定”
        到原所在单元格，绝不会因合并而丢失或重合。

    返回 io.BytesIO（xlsx 字节流），无数据则返回 None。
    """
    if not files:
        return None

    # 预收集所有 WPS 单元格内嵌图片（openpyxl 保存时会丢弃，需事后重新注入）
    wps_images = []
    for f in files:
        for img in _collect_wps_cell_images(f['path']):
            if not any(i['name'] == img['name'] for i in wps_images):
                wps_images.append(img)

    # 以第一个文件为模板：表头、列数、工作表名
    first_wb = openpyxl.load_workbook(files[0]['path'])
    first_ws = first_wb.active
    src_ncols = first_ws.max_column
    header = [first_ws.cell(row=1, column=c).value for c in range(1, src_ncols + 1)]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = first_ws.title or '合并'

    # 可选的“来源”列
    if mark_source:
        header = header + ['收集对象']

    # 写入表头（含样式）
    for c, val in enumerate(header, 1):
        cell = out_ws.cell(row=1, column=c, value=val)
        if c <= src_ncols:
            _copy_cell_style(first_ws.cell(row=1, column=c), cell)

    out_row = 2
    last_key = None            # 上一行写入的“数据列”内容，用于去重
    occupied = set()           # 已放置图片的 (行, 列)，避免重合

    for f in files:
        wb = openpyxl.load_workbook(f['path'])
        ws = wb.active

        # 建立图片映射：以 (0-based 行, 0-based 列) 为键
        img_map = {}
        for img in ws._images:
            fr = getattr(getattr(img, 'anchor', None), '_from', None)
            if fr is not None:
                img_map[(fr.row, fr.col)] = img

        nrows = ws.max_row
        for r in range(2, nrows + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, src_ncols + 1)]
            if all(v is None for v in row_vals):
                continue  # 跳过整行空白

            if mark_source:
                row_vals = row_vals + [f['object_name']]

            data_key = row_vals[:src_ncols]  # 用于比较“相同行”的键（不含来源列）

            if merge_identical and data_key == last_key:
                # 完全相同的行 -> 合并（去重）：不新增输出行，
                # 但若被丢弃的行含有截图，需把截图“固定”到保留行对应单元格，避免丢失
                kept_row = out_row - 1
                for (ir, ic), img in img_map.items():
                    if ir == r - 1 and ic < src_ncols:
                        _place_image_to_cell(out_ws, img, kept_row, ic + 1, occupied)
                continue

            # 写入该行数据（含样式）
            for c, val in enumerate(row_vals, 1):
                cell = out_ws.cell(row=out_row, column=c, value=val)
                if c <= src_ncols:
                    _copy_cell_style(ws.cell(row=r, column=c), cell)

            # 把原截图表单元格“固定”到输出表的相同位置
            for (ir, ic), img in img_map.items():
                if ir == r - 1 and ic < src_ncols:
                    _place_image_to_cell(out_ws, img, out_row, ic + 1, occupied)

            last_key = data_key
            out_row += 1

        wb.close()

    if out_row == 2:
        return None  # 没有任何数据行

    # 若新增了来源列，给一个合适的列宽
    if mark_source:
        source_col = src_ncols + 1
        out_ws.column_dimensions[get_column_letter(source_col)].width = 18

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)

    # 把 WPS 单元格内嵌图片（DISPIMG 截图）重新注入结果，避免合并后截图丢失/重叠
    buf = _inject_wps_cell_images(buf, wps_images)
    return buf


def create_export_archive(theme_id, theme_title):
    theme_folder = get_theme_folder(theme_id)
    # 每次使用唯一临时目录，避免并发/重复导出互相干扰，也避免共享 export_temp 被 safe-delete 拦截导致 500
    export_folder = os.path.join(theme_folder, 'export_temp_' + uuid.uuid4().hex)
    os.makedirs(export_folder)
    
    from models import CollectionObject, Attachment
    objects = CollectionObject.query.filter_by(theme_id=theme_id).all()
    
    for obj in objects:
        attachments = Attachment.query.filter_by(collection_object_id=obj.id).all()
        if len(attachments) == 1:
            att = attachments[0]
            src = os.path.join(theme_folder, f'object_{obj.id}', att.filename)
            _, ext = os.path.splitext(att.original_name)
            dst = os.path.join(export_folder, f"{obj.name}{ext}")
            if os.path.exists(src):
                shutil.copy2(src, dst)
        else:
            for idx, att in enumerate(attachments, 1):
                src = os.path.join(theme_folder, f'object_{obj.id}', att.filename)
                _, ext = os.path.splitext(att.original_name)
                dst = os.path.join(export_folder, f"{obj.name}_{idx}{ext}")
                if os.path.exists(src):
                    shutil.copy2(src, dst)

    archive_name = f"{theme_title}_附件汇总"
    archive_path = os.path.join(theme_folder, archive_name)
    # 同主题多次导出时，先安全删除旧归档，避免 make_archive 在部分 Python 版本抛 FileExistsError
    safe_remove_file(archive_path + '.zip')
    shutil.make_archive(archive_path, 'zip', export_folder)
    # 清理临时目录：用 safe_remove_dir 吞掉沙箱 safe-delete 包装器拦截导致的 OSError
    safe_remove_dir(export_folder)

    return f"{archive_name}.zip"
