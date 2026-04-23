import os
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from models import db, Admin, Announcement, AnnouncementAttachment, CollectionTheme, CollectionObject, Attachment, ThemeAttachment, beijing_now
from config import Config
from utils import (
    allowed_file, get_theme_folder, get_object_folder, get_announcement_folder,
    rename_uploaded_file, create_export_archive
)
import openpyxl

def register_routes(app):
    
    @app.template_filter('time_remaining')
    def time_remaining(deadline):
        now = beijing_now()
        if deadline < now:
            return '已截止'
        delta = deadline - now
        days = delta.days
        hours, remainder = divmod(delta.seconds, 3600)
        minutes = remainder // 60
        if days > 0:
            return f'{days}天{hours}小时'
        elif hours > 0:
            return f'{hours}小时{minutes}分钟'
        else:
            return f'{minutes}分钟'

    @app.template_filter('time_progress')
    def time_progress(deadline, created_at):
        now = beijing_now()
        if deadline < now:
            return 0
        total = (deadline - created_at).total_seconds()
        if total <= 0:
            return 0
        remaining = (deadline - now).total_seconds()
        if remaining <= 0:
            return 0
        progress = (remaining / total) * 100
        return max(0, min(100, int(progress)))

    @app.context_processor
    def inject_now():
        return {'now': beijing_now()}

    @app.route('/')
    def index():
        announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
        themes = CollectionTheme.query.filter_by(is_active=True).order_by(CollectionTheme.deadline.asc()).all()
        return render_template('index.html', announcements=announcements, themes=themes)

    @app.route('/theme/<int:theme_id>')
    def theme_detail(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        objects = CollectionObject.query.filter_by(theme_id=theme_id).all()
        completed = sum(1 for obj in objects if obj.is_completed)
        incomplete = [obj for obj in objects if not obj.is_completed]
        return render_template('theme_detail.html', theme=theme, objects=objects, 
                             completed=completed, incomplete=incomplete)

    @app.route('/upload/<int:object_id>', methods=['GET', 'POST'])
    def upload_page(object_id):
        collection_object = CollectionObject.query.get_or_404(object_id)
        theme = collection_object.theme
        
        if request.method == 'POST':
            if 'finish' in request.form:
                if collection_object.is_completed:
                    return redirect(url_for('index'))
                attachments = Attachment.query.filter_by(collection_object_id=object_id).all()
                if len(attachments) == 0:
                    flash('请先上传附件后再完成', 'error')
                    return redirect(url_for('upload_page', object_id=object_id))
                collection_object.is_completed = True
                collection_object.completed_at = beijing_now()
                db.session.commit()
                return redirect(url_for('index'))
            
            if 'file' in request.files:
                files = request.files.getlist('file')
                for file in files:
                    if file and file.filename and allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                        filename = secure_filename(file.filename)
                        stored_name = rename_uploaded_file(theme.id, collection_object.id, filename, filename)
                        folder = get_object_folder(theme.id, collection_object.id)
                        file_path = os.path.join(folder, stored_name)
                        file.save(file_path)
                        
                        attachment = Attachment(
                            filename=stored_name,
                            original_name=filename,
                            collection_object_id=collection_object.id
                        )
                        db.session.add(attachment)
                
                db.session.commit()
                return redirect(url_for('upload_page', object_id=object_id))
        
        attachments = Attachment.query.filter_by(collection_object_id=object_id).all()
        return render_template('upload.html', obj=collection_object, theme=theme, attachments=attachments)

    @app.route('/download/attachment/<int:attachment_id>')
    @login_required
    def download_attachment(attachment_id):
        attachment = Attachment.query.get_or_404(attachment_id)
        obj = attachment.collection_object
        file_path = os.path.join(get_object_folder(obj.theme.id, obj.id), attachment.filename)
        return send_file(file_path, as_attachment=True, download_name=attachment.original_name)

    @app.route('/announcement/download/<int:attachment_id>')
    def download_announcement_attachment(attachment_id):
        attachment = AnnouncementAttachment.query.get_or_404(attachment_id)
        file_path = os.path.join(get_announcement_folder(), attachment.filename)
        return send_file(file_path, as_attachment=True, download_name=attachment.original_name)

    @app.route('/theme/download/<int:attachment_id>')
    def download_theme_attachment(attachment_id):
        attachment = ThemeAttachment.query.get_or_404(attachment_id)
        file_path = os.path.join(get_theme_folder(attachment.theme_id), attachment.filename)
        return send_file(file_path, as_attachment=True, download_name=attachment.original_name)

    @app.route('/upload/delete-attachment/<int:attachment_id>', methods=['POST'])
    def delete_upload_attachment(attachment_id):
        attachment = Attachment.query.get_or_404(attachment_id)
        obj = attachment.collection_object
        file_path = os.path.join(get_object_folder(obj.theme.id, obj.id), attachment.filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        db.session.delete(attachment)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/admin/login', methods=['GET', 'POST'])
    def admin_login():
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            admin = Admin.query.filter_by(username=username).first()
            if admin and check_password_hash(admin.password_hash, password):
                login_user(admin)
                return redirect(url_for('admin_dashboard'))
            flash('用户名或密码错误', 'error')
        return render_template('admin_login.html')

    @app.route('/admin/logout')
    @login_required
    def admin_logout():
        logout_user()
        return redirect(url_for('index'))

    @app.route('/admin')
    @login_required
    def admin_dashboard():
        active_themes = CollectionTheme.query.filter_by(is_active=True).order_by(CollectionTheme.deadline.asc()).all()
        archived_themes = CollectionTheme.query.filter_by(is_active=False).order_by(CollectionTheme.created_at.desc()).all()
        announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
        return render_template('admin_dashboard.html', active_themes=active_themes, archived_themes=archived_themes, announcements=announcements)

    @app.route('/admin/change-password', methods=['GET', 'POST'])
    @login_required
    def change_password():
        if request.method == 'POST':
            old_password = request.form.get('old_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            if not current_user.check_password(old_password):
                flash('原密码错误', 'error')
                return redirect(url_for('change_password'))
            
            if new_password != confirm_password:
                flash('新密码与确认密码不一致', 'error')
                return redirect(url_for('change_password'))
            
            if len(new_password) < 6:
                flash('新密码长度不能少于6位', 'error')
                return redirect(url_for('change_password'))
            
            current_user.set_password(new_password)
            db.session.commit()
            flash('密码修改成功', 'success')
            return redirect(url_for('admin_dashboard'))
        
        return render_template('change_password.html')

    @app.route('/admin/theme/create', methods=['GET', 'POST'])
    @login_required
    def create_theme():
        if request.method == 'POST':
            title = request.form.get('title')
            description = request.form.get('description')
            announcement = request.form.get('announcement')
            deadline = request.form.get('deadline')
            collector_name = request.form.get('collector_name')
            collector_objects_text = request.form.get('collector_objects', '')
            
            theme = CollectionTheme(
                title=title,
                description=description,
                announcement=announcement,
                deadline=datetime.strptime(deadline, '%Y-%m-%dT%H:%M'),
                collector_name=collector_name
            )
            db.session.add(theme)
            db.session.commit()
            
            # 处理文本输入的收集对象
            if collector_objects_text.strip():
                object_names = collector_objects_text.strip().split('\n')
                for name in object_names:
                    name = name.strip()
                    if name:
                        obj = CollectionObject(name=name, theme_id=theme.id)
                        db.session.add(obj)
            
            # 处理Excel导入的收集对象
            if 'objects_excel_file' in request.files:
                file = request.files.get('objects_excel_file')
                if file and file.filename and file.filename.endswith(('.xls', '.xlsx')):
                    file.seek(0)
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        if row and len(row) > 0:
                            cell = row[0]
                            if cell is not None:
                                name = str(cell).strip()
                                if name:
                                    obj = CollectionObject(name=name, theme_id=theme.id)
                                    db.session.add(obj)
            
            if 'attachments' in request.files:
                files = request.files.getlist('attachments')
                for file in files:
                    if file and file.filename and allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                        original_name = file.filename
                        filename = secure_filename(original_name)
                        folder = get_theme_folder(theme.id)
                        counter = 1
                        while os.path.exists(os.path.join(folder, filename)):
                            name, ext = os.path.splitext(original_name)
                            filename = f"{name}_{counter}{ext}"
                            counter += 1
                        file_path = os.path.join(folder, filename)
                        file.save(file_path)
                        
                        att = ThemeAttachment(
                            filename=filename,
                            original_name=original_name,
                            theme_id=theme.id
                        )
                        db.session.add(att)
            db.session.commit()
            
            flash('收集主题创建成功', 'success')
            return redirect(url_for('admin_dashboard'))
        return render_template('theme_create.html')

    @app.route('/admin/theme/<int:theme_id>/objects', methods=['GET', 'POST'])
    @login_required
    def manage_theme_objects(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        
        if request.method == 'POST':
            if 'add_object' in request.form:
                name = request.form.get('object_name')
                obj = CollectionObject(name=name, theme_id=theme_id)
                db.session.add(obj)
                db.session.commit()
                flash('收集对象添加成功', 'success')
            
            elif 'import_excel' in request.form:
                if 'excel_file' in request.files:
                    file = request.files.get('excel_file')
                    if file and file.filename.endswith(('.xls', '.xlsx')):
                        wb = openpyxl.load_workbook(file)
                        ws = wb.active
                        imported_count = 0
                        for row in ws.iter_rows(values_only=True):
                            if row and len(row) > 0:
                                cell = row[0]
                                if cell is not None:
                                    name = str(cell).strip()
                                    if name:
                                        existing = CollectionObject.query.filter_by(name=name, theme_id=theme_id).first()
                                        if not existing:
                                            obj = CollectionObject(name=name, theme_id=theme_id)
                                            db.session.add(obj)
                                            imported_count += 1
                        db.session.commit()
                        flash(f'Excel导入成功，共导入 {imported_count} 条数据', 'success')
        
        objects = CollectionObject.query.filter_by(theme_id=theme_id).all()
        completed = sum(1 for obj in objects if obj.is_completed)
        incomplete = [obj for obj in objects if not obj.is_completed]
        
        return render_template('theme_objects.html', theme=theme, objects=objects, 
                             completed=completed, incomplete=incomplete)

    @app.route('/admin/theme/<int:theme_id>/delete', methods=['POST'])
    @login_required
    def delete_theme(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        theme_folder = get_theme_folder(theme_id)
        if os.path.exists(theme_folder):
            import shutil
            shutil.rmtree(theme_folder)
        db.session.delete(theme)
        db.session.commit()
        flash('主题已删除', 'success')
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/object/<int:object_id>/delete', methods=['POST'])
    @login_required
    def delete_object(object_id):
        obj = CollectionObject.query.get_or_404(object_id)
        db.session.delete(obj)
        db.session.commit()
        flash('收集对象已删除', 'success')
        return redirect(url_for('manage_theme_objects', theme_id=obj.theme_id))

    @app.route('/admin/object/<int:object_id>/reset', methods=['POST'])
    @login_required
    def reset_object_upload(object_id):
        obj = CollectionObject.query.get_or_404(object_id)
        attachments = Attachment.query.filter_by(collection_object_id=object_id).all()
        for att in attachments:
            file_path = os.path.join(get_object_folder(obj.theme.id, obj.id), att.filename)
            if os.path.exists(file_path):
                os.remove(file_path)
            db.session.delete(att)
        obj.is_completed = False
        obj.completed_at = None
        db.session.commit()
        flash('上传已重置，可以重新上传', 'success')
        return redirect(url_for('manage_theme_objects', theme_id=obj.theme_id))

    @app.route('/admin/attachment/<int:attachment_id>/delete', methods=['POST'])
    @login_required
    def delete_attachment(attachment_id):
        att = Attachment.query.get_or_404(attachment_id)
        obj = att.collection_object
        file_path = os.path.join(get_object_folder(obj.theme.id, obj.id), att.filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        db.session.delete(att)
        db.session.commit()
        flash('附件已删除', 'success')
        if request.referrer and 'object_id' in request.referrer:
            return redirect(url_for('manage_theme_objects', theme_id=obj.theme_id))
        return redirect(url_for('manage_theme_objects', theme_id=obj.theme_id))

    @app.route('/admin/object/<int:object_id>/download-attachments')
    @login_required
    def download_object_attachments(object_id):
        obj = CollectionObject.query.get_or_404(object_id)
        attachments = Attachment.query.filter_by(collection_object_id=object_id).all()
        if not attachments:
            flash('该对象没有附件', 'error')
            return redirect(url_for('manage_theme_objects', theme_id=obj.theme_id))
        
        if len(attachments) == 1:
            att = attachments[0]
            file_path = os.path.join(get_object_folder(obj.theme.id, obj.id), att.filename)
            ext = os.path.splitext(att.original_name)[1]
            download_name = f"{obj.name}{ext}"
            return send_file(file_path, as_attachment=True, download_name=download_name)
        else:
            import io
            import zipfile
            memory_file = io.BytesIO()
            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                for att in attachments:
                    file_path = os.path.join(get_object_folder(obj.theme.id, obj.id), att.filename)
                    if os.path.exists(file_path):
                        zf.write(file_path, f"{obj.name}_{att.original_name}")
            memory_file.seek(0)
            return send_file(memory_file, as_attachment=True, download_name=f"{obj.name}_附件.zip")

    @app.route('/admin/theme/<int:theme_id>/export')
    @login_required
    def export_theme_attachments(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        archive_name = create_export_archive(theme_id, theme.title)
        archive_path = os.path.join(get_theme_folder(theme_id), archive_name)
        return send_file(archive_path, as_attachment=True)

    @app.route('/admin/announcement/create', methods=['GET', 'POST'])
    @login_required
    def create_announcement():
        if request.method == 'POST':
            title = request.form.get('title')
            content = request.form.get('content')
            
            announcement = Announcement(title=title, content=content)
            db.session.add(announcement)
            db.session.commit()
            
            if 'attachments' in request.files:
                files = request.files.getlist('attachments')
                for file in files:
                    if file and file.filename and allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                        original_name = file.filename
                        filename = secure_filename(original_name)
                        folder = get_announcement_folder()
                        counter = 1
                        while os.path.exists(os.path.join(folder, filename)):
                            name, ext = os.path.splitext(original_name)
                            filename = f"{name}_{counter}{ext}"
                            counter += 1
                        file_path = os.path.join(folder, filename)
                        file.save(file_path)
                        
                        att = AnnouncementAttachment(
                            filename=filename,
                            original_name=original_name,
                            announcement_id=announcement.id
                        )
                        db.session.add(att)
            
            db.session.commit()
            flash('公告创建成功', 'success')
            return redirect(url_for('admin_dashboard'))
        return render_template('announcement_create.html')

    @app.route('/admin/announcement/<int:announcement_id>/delete', methods=['POST'])
    @login_required
    def delete_announcement(announcement_id):
        announcement = Announcement.query.get_or_404(announcement_id)
        for att in announcement.attachments:
            file_path = os.path.join(get_announcement_folder(), att.filename)
            if os.path.exists(file_path):
                os.remove(file_path)
        db.session.delete(announcement)
        db.session.commit()
        flash('公告已删除', 'success')
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/announcement/<int:announcement_id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_announcement(announcement_id):
        announcement = Announcement.query.get_or_404(announcement_id)
        if request.method == 'POST':
            announcement.title = request.form.get('title')
            announcement.content = request.form.get('content')
            db.session.commit()
            
            if 'attachments' in request.files:
                files = request.files.getlist('attachments')
                for file in files:
                    if file and file.filename and allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                        original_name = file.filename
                        filename = secure_filename(original_name)
                        folder = get_announcement_folder()
                        counter = 1
                        while os.path.exists(os.path.join(folder, filename)):
                            name, ext = os.path.splitext(original_name)
                            filename = f"{name}_{counter}{ext}"
                            counter += 1
                        file_path = os.path.join(folder, filename)
                        file.save(file_path)
                        
                        att = AnnouncementAttachment(
                            filename=filename,
                            original_name=original_name,
                            announcement_id=announcement.id
                        )
                        db.session.add(att)
            db.session.commit()
            flash('公告已更新', 'success')
            return redirect(url_for('admin_dashboard'))
        return render_template('announcement_edit.html', announcement=announcement)

    @app.route('/admin/announcement/attachment/<int:attachment_id>/delete', methods=['GET', 'POST'])
    @login_required
    def delete_announcement_attachment(attachment_id):
        attachment = AnnouncementAttachment.query.get_or_404(attachment_id)
        announcement_id = attachment.announcement_id
        file_path = os.path.join(get_announcement_folder(), attachment.filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        db.session.delete(attachment)
        db.session.commit()
        flash('附件已删除', 'success')
        return redirect(url_for('edit_announcement', announcement_id=announcement_id))

    @app.route('/admin/theme/<int:theme_id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_theme(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        if request.method == 'POST':
            theme.title = request.form.get('title')
            theme.description = request.form.get('description')
            theme.announcement = request.form.get('announcement')
            theme.deadline = datetime.strptime(request.form.get('deadline'), '%Y-%m-%dT%H:%M')
            theme.collector_name = request.form.get('collector_name')
            
            if 'attachments' in request.files:
                files = request.files.getlist('attachments')
                for file in files:
                    if file and file.filename and allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                        original_name = file.filename
                        filename = secure_filename(original_name)
                        folder = get_theme_folder(theme.id)
                        counter = 1
                        while os.path.exists(os.path.join(folder, filename)):
                            name, ext = os.path.splitext(original_name)
                            filename = f"{name}_{counter}{ext}"
                            counter += 1
                        file_path = os.path.join(folder, filename)
                        file.save(file_path)
                        
                        att = ThemeAttachment(
                            filename=filename,
                            original_name=original_name,
                            theme_id=theme.id
                        )
                        db.session.add(att)
            
            db.session.commit()
            flash('主题已更新', 'success')
            return redirect(url_for('admin_dashboard'))
        return render_template('theme_edit.html', theme=theme)

    @app.route('/admin/theme/<int:theme_id>/toggle', methods=['POST'])
    @login_required
    def toggle_theme_status(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        theme.is_active = not theme.is_active
        db.session.commit()
        return jsonify({'success': True, 'is_active': theme.is_active})

    @app.route('/admin/theme/<int:theme_id>/archive', methods=['POST'])
    @login_required
    def archive_theme(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        theme.is_active = False
        db.session.commit()
        flash('主题已归档', 'success')
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/theme/<int:theme_id>/restore', methods=['POST'])
    @login_required
    def restore_theme(theme_id):
        theme = CollectionTheme.query.get_or_404(theme_id)
        theme.is_active = True
        db.session.commit()
        flash('主题已恢复', 'success')
        return redirect(url_for('manage_theme_objects', theme_id=theme_id))

    @app.errorhandler(404)
    def not_found(e):
        return render_template('404.html'), 404
